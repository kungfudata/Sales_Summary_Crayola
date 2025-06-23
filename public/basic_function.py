# -*- coding:utf-8 -*-
import os
from email.mime.text import MIMEText
from email.utils import formataddr
from email.header import Header
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles import PatternFill, Border, Side
# from DBUtils.PooledDB import PooledDB

import pymysql
import random

from public import config
import requests
import hashlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warehouse_no_List = ['crayola-2', 'crayola-3-1', 'crayola-4', 'crayola-3', 'crayola-3-2']


class DataBase_baidu:
    def __init__(self, db_name="store_report", host="180.76.60.179", port=3306):
        self.host = host
        self.port = port
        self.user = config.user
        self.passwd = config.passwd
        self.db = db_name
        self.conn = None
        self.cur = None
        self.charset = "utf8"
        self.connect()

    def connect(self):
        try:
            conn = pymysql.connect(
                host=self.host,
                port=self.port,
                user=self.user,
                passwd=self.passwd,
                db=self.db,
                charset=self.charset
            )
            cur = conn.cursor()
            self.conn = conn
            self.cur = cur
        except pymysql.MySQLError as e:
            print(f"Error connecting to database: {e}")

    def execute_query(self, query, params=None):
        try:
            self.cur.execute(query, params)
            self.conn.commit()
        except pymysql.MySQLError as e:
            print(f"Error executing query: {e}")
            self.conn.rollback()

    def commit(self):
        try:
            self.conn.commit()
        except pymysql.MySQLError as e:
            print(f"Error committing transaction: {e}")
            self.conn.rollback()

    def fetchall(self):
        return self.cur.fetchall()

    def description(self):
        return self.cur.description

    def fetchone(self):
        return self.cur.fetchone()

    def close(self):
        if self.cur:
            self.cur.close()
        if self.conn:
            self.conn.close()
        print("Database connection closed")

    def get_account_info(self, account_name):
        try:
            query = "SELECT sid, appkey, target_app_key, appsecret FROM WDT_integration.wdt_account_info WHERE sid = %s"
            self.cur.execute(query, (account_name,))
            result = self.cur.fetchone()
            if result:
                sid, appkey, target_app_key, appsecret = result
                return sid, appkey, target_app_key, appsecret
            else:
                print("account not found")
                return None
        except pymysql.MySQLError as e:
            print(f"Error executing query: {e}")
            return None


class DataBase_baidu_store_report:
    def __init__(self, db_name="store_report", host="180.76.60.179", port=3306):
        self.host = host
        self.port = port
        self.user = config.user
        self.passwd = config.passwd
        self.db = db_name
        self.conn = None
        self.cur = None
        self.charset = "utf8"
        self.connect()

    def connect(self):
        try:
            conn = pymysql.connect(
                host=self.host,
                port=self.port,
                user=self.user,
                passwd=self.passwd,
                db=self.db,
                charset=self.charset
            )
            cur = conn.cursor()
            self.conn = conn
            self.cur = cur
        except pymysql.MySQLError as e:
            print(f"Error connecting to database: {e}")

    def execute_query(self, query, params=None):
        try:
            self.cur.execute(query, params)
            self.conn.commit()
        except pymysql.MySQLError as e:
            print(f"Error executing query: {e}")
            self.conn.rollback()

    def commit(self):
        try:
            self.conn.commit()
        except pymysql.MySQLError as e:
            print(f"Error committing transaction: {e}")
            self.conn.rollback()


class RandomUserAgent:
    def __init__(self):
        self.user_agent_list = [
            "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.62 Safari/537.36",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
            "Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10.5; en-US; rv:1.9.2.15) Gecko/20110303 Firefox/3.6.15",
        ]

    def get_random_userAgent(self):
        user_agent = random.choice(self.user_agent_list)
        return user_agent


class MailGunner:
    def __init__(self):
        self.PrivateApiKey = config.mailgun_PrivateApiKey
        self.PublicValidationKey = config.mailgun_PublicValidationKey
        self.HTTPWebhookSigningKey = config.mailgun_HTTPWebhookSigningKey
        self.domain = config.mailgun_domain

    def send_WDT_script_error_alert(self, sid, script_name, e):
        recipients = ['techreports@kungfudata.com']
        subject = "!!!卖家账号" + sid + "旺店通数据导入脚本报错。"
        emailtext = '卖家账号为：' + sid + '，在导入旺店通数据到百度数据库时，' + script_name + '报错。 报错信息为：' + str(e)
        return requests.post(
            "https://api.mailgun.net/v3/{}/messages".format(self.domain),
            auth=("api", self.PrivateApiKey),
            data={"from": "KF脚本运行错误 <noreply@{}>".format(self.domain),
                  "to": recipients,
                  "subject": subject,
                  "text": emailtext})

    def send_Tplus_error_alert(self, script_name, e):
        recipients = ['techreports@kungfudata.com']
        subject = "!!!" + "同步数据库数据到Tplus时脚本报错。"
        emailtext = '同步数据库数据到Tplus时，' + script_name + '报错。 报错信息为：' + str(e)
        return requests.post(
            "https://api.mailgun.net/v3/{}/messages".format(self.domain),
            auth=("api", self.PrivateApiKey),
            data={"from": "KF脚本运行错误 <noreply@{}>".format(self.domain),
                  "to": recipients,
                  "subject": subject,
                  "text": emailtext})

    def send_Tplus_token_error_alert(self, script_name, e):
        recipients = ['techreports@kungfudata.com', 'pengyue@kungfudata.com']
        subject = "!!!" + "Tplus获取token出错。"
        emailtext = 'Tplus token获取失败，请马上检查！' + script_name + '报错。 报错信息为：' + str(e)
        return requests.post(
            "https://api.mailgun.net/v3/{}/messages".format(self.domain),
            auth=("api", self.PrivateApiKey),
            data={"from": "KF脚本运行错误 <noreply@{}>".format(self.domain),
                  "to": recipients,
                  "subject": subject,
                  "text": emailtext})

    def send_no_unit_name(self, blank_unit_List):
        recipients = ['techreports@kungfudata.com', 'pengyue@kungfudata.com']
        subject = "!!!" + "导入旺店通货品档案到百度数据库时，有的产品没有基本单位。"
        emailtext = '导入旺店通货品档案到数据库时，有的产品没有基本单位，请联系运营，让他们及时为货品添加单位。缺少基本单位的产品的卖家账号和货品编码分别为：\n\n'
        for sid, spec_no in blank_unit_List:
            emailtext += f"卖家账号 = {sid}, 货品编码 = {spec_no}\n"

        return requests.post(
            "https://api.mailgun.net/v3/{}/messages".format(self.domain),
            auth=("api", self.PrivateApiKey),
            data={"from": "KF脚本运行错误 <noreply@{}>".format(self.domain),
                  "to": recipients,
                  "subject": subject,
                  "text": emailtext})

    def send_Tplus_Products_need_to_be_manually_do(self, script_name, sql):
        recipients = ['techreports@kungfudata.com', 'pengyue@kungfudata.com']
        subject = "!!!" + "旺店通以下货品已成功同步到T+，但是记录该货品已经成功同步的sql语句没有执行成功，请手动执行并检查原因"
        emailtext = '脚本' + str(script_name) + '旺店通以下货品已成功同步到T+，但是记录该货品已经成功同步的sql语句没有执行成功，请手动执行并检查原因' + str(sql)
        return requests.post(
            "https://api.mailgun.net/v3/{}/messages".format(self.domain),
            auth=("api", self.PrivateApiKey),
            data={"from": "KF脚本运行错误 <noreply@{}>".format(self.domain),
                  "to": recipients,
                  "subject": subject,
                  "text": emailtext})

    def send_Tplus_QiTaChuKuDan_Aduit(self, order_num):
        recipients = ['techreports@kungfudata.com']
        subject = "!!!" + "请登录T+审核其他出入库单"
        emailtext = f'旺店通单据已导入，但是单据需手动审核，请登录T+在其他出入库单中，审核旺店通出入库单号为【{order_num}】的单据'
        return requests.post(
            "https://api.mailgun.net/v3/{}/messages".format(self.domain),
            auth=("api", self.PrivateApiKey),
            data={"from": "KF脚本提醒 <noreply@{}>".format(self.domain),
                  "to": recipients,
                  "subject": subject,
                  "text": emailtext})


class PlatformSign:
    def __init__(self):
        pass

    def get_QiMen_sign(self, params, appsecret):
        # 生成sign
        # https://open.taobao.com/doc.htm?docId=101617&docType=1
        # https://open.wangdian.cn/open/guide?path=guide_signsf

        # 排序参数
        params = sorted(params.items(), key=lambda x: x[0])

        # 构造签名字符串
        sign_List = []
        for single in params:
            key_name = single[0]
            value_name = single[1]
            string_final = str(key_name) + str(value_name)

            sign_List.append(string_final)
            # body_List.append(str(value_name))

        f = lambda x: x.strip()
        content_list = [f(x) for x in sign_List]
        sign_string = ''.join(content_list)

        # 将appsecret添加到签名字符串的前后
        # body_string = json.dumps(params)
        # sign_string = appsecret + sign_string + appsecret
        sign_string = appsecret + sign_string + appsecret

        # 计算md5签名
        sign_string = sign_string.encode('UTF-8')
        # print(sign_string)
        m = hashlib.md5()
        m.update(sign_string)
        sign = m.hexdigest()
        sign = str(sign).upper()
        # print(sign)
        return sign

    def get_WDT_sign(self, params, app_secret):
        appsecret = app_secret
        # 生成sign
        # https://open.taobao.com/doc.htm?docId=101617&docType=1
        # https://open.wangdian.cn/open/guide?path=guide_signsf
        params = sorted(params.items(), key=lambda x: x[0])

        sign_List = []

        for single in params:
            key_name = single[0]
            value_name = single[1]
            string10 = len(str(key_name))
            if len(str(string10)) > 2:
                string10 = str(string10)[:2]
            elif len(str(string10)) < 2:
                string10 = '0' + str(string10)
            else:
                string10 = str(string10)

            string10 = string10 + '-' + str(key_name)

            string11 = len(str(value_name))
            if len(str(string11)) >= 4:
                string11 = str(string11)
            elif len(str(string11)) < 4:
                numbers_of_zero = 4 - len(str(string11))
                # print(numbers_of_zero)
                if numbers_of_zero == 1:
                    string11 = '0' + str(string11)
                if numbers_of_zero == 2:
                    string11 = '00' + str(string11)
                if numbers_of_zero == 3:
                    string11 = '000' + str(string11)
                if numbers_of_zero == 4:
                    string11 = '0000'

            string11 = string11 + '-' + str(value_name)

            string_final = string10 + ':' + string11

            sign_List.append(string_final)

        f = lambda x: x.strip()
        content_list = [f(x) for x in sign_List]
        sign_string = ';'.join(content_list)
        sign_string = sign_string + appsecret
        # print(sign_string)
        m = hashlib.md5()
        m.update(sign_string)
        sign = m.hexdigest()
        # print(sign)
        return sign


# class ExcelSaver:
#     def __init__(self, file_name):
#         """
#         初始化ExcelSaver类
#
#         :param file_name: 要保存的Excel文件名
#         """
#         self.file_name = file_name
#         self.workbook = Workbook()
#         self.sheet = self.workbook.active
#         self.sheet.title = "Sheet1"  # 设置默认工作表名称
#
#     def save_dataframe(self, df, start_row, start_col):
#         """
#         将DataFrame保存到Excel的指定位置
#
#         :param df: 要保存的DataFrame
#         :param start_row: 数据框的起始行
#         :param start_col: 数据框的起始列
#         """
#         # 将DataFrame转化为Excel行
#         for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
#             for c_idx, value in enumerate(row, start=start_col):
#                 self.sheet.cell(row=r_idx, column=c_idx, value=value)
#
#     def save_text(self, text, start_row, start_col):
#         """
#         将字符串保存到Excel的指定位置
#
#         :param text: 要保存的字符串
#         :param start_row: 字符串的起始行
#         :param start_col: 字符串的起始列
#         """
#         self.sheet.cell(row=start_row, column=start_col, value=text)
#
#     def fill_cell_red(self, row, col):
#         """
#         将指定单元格填充为红色
#
#         :param row: 行号
#         :param col: 列号
#         """
#         red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
#         cell = self.sheet.cell(row=row, column=col)
#         cell.fill = red_fill
#
#     def save(self):
#         """
#         保存Excel文件
#         """
#         self.workbook.save(self.file_name)
class ExcelSaver:
    def __init__(self, file_name):
        """
        初始化ExcelSaver类

        :param file_name: 要保存的Excel文件名
        """
        self.file_name = file_name
        self.workbook = Workbook()
        # 默认删除自动创建的Sheet
        self.workbook.remove(self.workbook.active)
        self.sheets = {}  # 用于存储各个sheet的名称和内容

    def create_sheet(self, sheet_name):
        """
        创建一个新的sheet并指定名称

        :param sheet_name: 新sheet的名称
        """
        if sheet_name not in self.sheets:
            self.sheets[sheet_name] = self.workbook.create_sheet(sheet_name)

    def save_dataframe(self, df, sheet_name, start_row, start_col):
        """
        将DataFrame保存到指定的Sheet和位置

        :param df: 要保存的DataFrame
        :param sheet_name: 要保存的sheet名称
        :param start_row: 数据框的起始行
        :param start_col: 数据框的起始列
        """
        self.create_sheet(sheet_name)  # 确保目标sheet存在
        sheet = self.sheets[sheet_name]

        # 将DataFrame转化为Excel行
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    def save_text(self, text, sheet_name, start_row, start_col):
        """
        将字符串保存到指定的Sheet和位置

        :param text: 要保存的字符串
        :param sheet_name: 要保存的sheet名称
        :param start_row: 字符串的起始行
        :param start_col: 字符串的起始列
        """
        self.create_sheet(sheet_name)  # 确保目标sheet存在
        sheet = self.sheets[sheet_name]
        sheet.cell(row=start_row, column=start_col, value=text)

    def fill_cell_red(self, sheet_name, row, col):
        """
        将指定sheet的指定单元格填充为红色

        :param sheet_name: 要填充颜色的sheet名称
        :param row: 行号
        :param col: 列号
        """
        self.create_sheet(sheet_name)  # 确保目标sheet存在
        sheet = self.sheets[sheet_name]
        red_fill = PatternFill(start_color="D2E4FC", end_color="D2E4FC", fill_type="solid")
        cell = sheet.cell(row=row, column=col)
        cell.fill = red_fill

    def add_bold_border(self, sheet_name, start_row, start_col, end_row, end_col):
        """
        给指定的单元格区域加粗外边框

        :param sheet_name: 要操作的sheet名称
        :param start_row: 起始行号
        :param start_col: 起始列号
        :param end_row: 结束行号
        :param end_col: 结束列号
        """
        self.create_sheet(sheet_name)  # 确保目标sheet存在
        sheet = self.sheets[sheet_name]

        # 定义加粗边框样式
        thick_border = Side(border_style="thick", color="000000")

        # 更新边框的辅助函数
        def update_border(cell, top=None, left=None, right=None, bottom=None):
            existing_border = cell.border
            cell.border = Border(
                top=top if top else existing_border.top,
                left=left if left else existing_border.left,
                right=right if right else existing_border.right,
                bottom=bottom if bottom else existing_border.bottom,
            )

        # 为边框区域添加外边框
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                if row == start_row:  # 上边框
                    update_border(cell, top=thick_border)
                if row == end_row:  # 下边框
                    update_border(cell, bottom=thick_border)
                if col == start_col:  # 左边框
                    update_border(cell, left=thick_border)
                if col == end_col:  # 右边框
                    update_border(cell, right=thick_border)

    def delete_crayola_files(self, file_path):
        """
        删除file_path下包含'Sales status of Crayola'的xlsx文件

        :param file_path: 要搜索的文件目录路径
        """
        if not os.path.exists(file_path):
            print(f"路径 {file_path} 不存在！")
            return

        for file_name in os.listdir(file_path):
            if "Sales status of Crayola" in file_name and file_name.endswith(".xlsx"):
                full_path = os.path.join(file_path, file_name)
                try:
                    os.remove(full_path)
                    print(f"已删除文件: {full_path}")
                except Exception as e:
                    print(f"无法删除文件: {full_path}，错误: {e}")

    def save(self):
        """
        保存Excel文件
        """
        self.workbook.save(self.file_name)


class EmailSender:
    def __init__(self):
        pass

    def send_report_to_people(self, store_name, file_path, email_list, start_time, end_time):
        # 创建MIMEMultipart对象
        msg = MIMEMultipart()
        # 收件人、发件人、邮件标题
        # sender = '1551848014@qq.com'
        sender = '1183780564@qq.com'
        msg['From'] = formataddr(["Pengyue", sender])
        print(email_list, type(email_list))
        msg['To'] = ', '.join(email_list)
        msg['Cc'] = 'techreports@kungfudata.com'
        msg['Subject'] = Header(f'{store_name} Sales Performance', 'utf-8')

        # 正文部分
        textmessage = MIMEText(
            '<p>Here is the ' + f' {store_name} sales performance from {start_time} to {end_time}.Please check the attachement.</p>' +
            '<br>' +
            '<br>' +
            '<p>Best Regards,</p>' +
            '<p>Pengyue</p>',
            'html', 'utf-8')
        msg.attach(textmessage)

        # 创建一个MIMEBase对象并设置相应的头信息
        part = MIMEBase('application', 'octet-stream')
        with open(file_path, 'rb') as attachment:
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        file_name = os.path.basename(file_path)
        part.add_header('Content-Disposition', f"attachment; filename={file_name}")

        # 将附件添加到消息对象中
        msg.attach(part)

        smtpObj = smtplib.SMTP('smtp.qq.com', 587)
        smtpObj.ehlo()
        smtpObj.starttls()
        # smtpObj.login('1551848014@qq.com', 'untjokeusbhnhffd')
        smtpObj.login('1183780564@qq.com', 'mcimgmsgeahpbacd')
        smtpObj.sendmail(msg['From'], email_list, msg.as_string())
        smtpObj.quit()
